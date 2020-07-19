# Written by: Elin Liu June 2020
import openpyxl
import os
from selenium import webdriver

# Load the data
os.chdir('C:\\Users\\Elin Liu\\Desktop\\Elins Stuff\\School\\Online courses\\Python Intro\\Covid Lit')
workbook_load = openpyxl.load_workbook('LitCovid_21364_Jun12_2020_authInfo6.xlsx')
sheet = workbook_load['LitCovid_21364_Jun12_2020_Text']

# Initialize variables 
publications_full = []
name = []
journal = []

# Note: 53266 cells in sheet
for i in range(1,53267):
    if(sheet.cell(row=i,column=1).value!= None):
        publications_full.append(str(sheet.cell(row=i,column=1).value))

# Clean data to find article name
for publication in publications_full:
    info = publication.split("\"")
    art_name = ""
    for i in range (1,len(info)-1):
        art_name = art_name + info[i]
    name.append(art_name)
    journal.append(info[len(info)-1])

articles = len(name)
print(articles)
print("completed loading data")

# Export the data, separate name and journal 
workbook_load.create_sheet(title='articles')
sheet1 = workbook_load['articles']
sheet1['A1'] = 'Name'
sheet1['B1'] = 'Journal'

for i in range(articles):
   sheet1.cell(row=i+2,column=1).value = name[i]
   sheet1.cell(row=i+2,column=2).value = journal[i]

# Save article Name and journal
workbook_load.save('LitCovid_21364_Jun12_2020_TextUpdate2.xlsx')

# Find all unique journals and frequencies 
unique_articles = {}
for i in range (articles):
   jname = journal[i]
   exists = jname in unique_articles.keys()
   if(not exists):
       unique_articles[journal[i]]= 1
   else:
       counter = int (unique_articles.get(jname))+ 1
       update_item = {jname: counter}
       unique_articles.update(update_item)

# Save unique journals and frequency
workbook_load.create_sheet(title='unique journals')
sheet1 = workbook_load['unique journals']

i = 0
for item in unique_articles.items():
   sheet1.cell(row=i+1,column=1).value = item[0]
   sheet1.cell(row=i+1,column=2).value = item[1]
   i = i+1

workbook_load.save('LitCovid_21364_Jun12_2020_TextUpdate3.xlsx')

# Find information about the first author on pubmed 
first_auth_info = {}

browser = webdriver.Firefox()

for i in range(0,21635):
    article = name[i]
    workbook_load = openpyxl.load_workbook('LitCovid_21364_Jun12_2020_authInfo6.xlsx')
    sheet1 = workbook_load['author_info']
    workbook_load.save('LitCovid_21364_Jun12_2020_authInfo6.xlsx')
    
    # Search article name on pubmed
    browser.get('https://pubmed.ncbi.nlm.nih.gov/')
    searchBar = browser.find_element_by_css_selector('#id_term')
    searchBar.send_keys('"' + article + '"')
    searchButton = browser.find_element_by_css_selector('button.search-btn')
    searchButton.click()
    try:
        # Click first link if applicable
        firstLink = browser.find_element_by_css_selector('a.labs-docsum-title[data-ga-action="1"')
        firstLink.click()
    except:
        print("")
    try:
        # Toggle author info if applicable 
        expand = browser.find_element_by_css_selector('#toggle-authors')
        expand.click()
        
        # Only record information from first author 
        first_aff = browser.find_element_by_css_selector('div.affiliations:nth-child(1)>ul:nth-child(2)>li:nth-child(1)')
        aff_info = first_aff.text
    except:
        aff_info = "No Affiliation Info"
    
    # Save author information
    first_auth_info[article] = aff_info
    sheet1.cell(row=i+1,column=1).value = article
    sheet1.cell(row=i+1,column=2).value = aff_info
    workbook_load.save('LitCovid_21364_Jun12_2020_authInfo6.xlsx')
    print(i+1)
    i = i+1
browser.quit()

print("done finding author info")

# workbook_load.create_sheet(title='author_info')
# sheet1 = workbook_load['author_info']

# i = 0
# for item in first_auth_info.items():
#    sheet1.cell(row=i+1,column=1).value = item[0]
#    sheet1.cell(row=i+1,column=2).value = item[1]
#    i = i+1

workbook_load.save('LitCovid_21364_Jun12_2020_authInfo6.xlsx')

    
    



    
